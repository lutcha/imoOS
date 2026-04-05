"""
Budget renderers — Exportação de orçamentos para PDF e Excel.
"""
import io
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.budget.models import BudgetItem


def format_currency(value):
    """Formata valor em CVE."""
    if value is None:
        return "0,00"
    return f"{value:,.2f} CVE".replace(",", "X").replace(".", ",").replace("X", ".")


class BudgetPDFRenderer:
    """Renderer para exportação de orçamento em PDF."""
    
    def render(self, budget, include_notes=True):
        """Renderiza orçamento como PDF."""
        try:
            from weasyprint import HTML, CSS
            USE_WEASYPRINT = True
        except ImportError:
            USE_WEASYPRINT = False
        
        # Preparar contexto
        items_by_category = self._group_items_by_category(budget)
        
        context = {
            'budget': budget,
            'items_by_category': items_by_category,
            'include_notes': include_notes,
            'generated_at': datetime.now(),
            'company_name': 'ImoOS',
        }
        
        # Renderizar HTML
        html_string = render_to_string('budget/pdf_budget.html', context)
        
        if USE_WEASYPRINT:
            # Usar WeasyPrint para gerar PDF
            html = HTML(string=html_string)
            pdf = html.write_pdf()
            
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="orcamento_{budget.name.replace(" ", "_")}.pdf"'
            return response
        else:
            # Fallback: retornar HTML
            response = HttpResponse(html_string, content_type='text/html')
            response['Content-Disposition'] = f'attachment; filename="orcamento_{budget.name.replace(" ", "_")}.html"'
            return response
    
    def _group_items_by_category(self, budget):
        """Agrupa itens do orçamento por categoria."""
        from collections import defaultdict
        
        groups = defaultdict(list)
        for item in budget.items.all():
            category_name = item.get_category_display()
            category_code = item.category
            groups[category_code].append({
                'category_name': category_name,
                'item': item
            })
        
        # Ordenar por código de categoria
        return dict(sorted(groups.items()))


class BudgetExcelRenderer:
    """Renderer para exportação de orçamento em Excel."""
    
    def render(self, budget, include_notes=True):
        """Renderiza orçamento como Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Orçamento"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_font = Font(bold=True, size=11)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = budget.name
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Info do orçamento
        ws['A3'] = 'Descrição:'
        ws['B3'] = budget.description or '-'
        ws['A4'] = 'Ilha:'
        ws['B4'] = budget.get_island_display()
        ws['A5'] = 'Status:'
        ws['B5'] = budget.get_status_display()
        ws['A6'] = 'Contingência:'
        ws['B6'] = f"{budget.contingency_pct}%"
        
        # Cabeçalhos
        header_row = 8
        headers = ['Ord', 'Item', 'Unidade', 'Qtd', 'Preço Unit.', 'Total']
        if include_notes:
            headers.append('Notas')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Dados
        current_row = header_row + 1
        current_category = None
        
        items = budget.items.all().order_by('line_number')
        
        for item in items:
            # Verificar se mudou de categoria
            category_name = item.get_category_display()
            
            if category_name != current_category:
                current_category = category_name
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                cat_cell = ws.cell(row=current_row, column=1, value=category_name)
                cat_cell.fill = subheader_fill
                cat_cell.font = subheader_font
                cat_cell.border = thin_border
                current_row += 1
            
            # Dados do item
            data = [
                item.line_number,
                item.description,
                item.get_unit_display(),
                float(item.quantity),
                float(item.unit_price),
                float(item.total),
            ]
            if include_notes:
                data.append(item.notes)
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                
                # Formatação de número
                if col in [4, 5, 6]:  # Qtd, Preço, Total
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # Totais
        totals_row = current_row + 1
        ws.cell(row=totals_row, column=1, value='Subtotal:').font = total_font
        ws.cell(row=totals_row, column=6, value=float(budget.subtotal)).font = total_font
        ws.cell(row=totals_row, column=6).number_format = '#,##0.00 CVE'
        ws.cell(row=totals_row, column=6).fill = total_fill
        
        ws.cell(row=totals_row + 1, column=1, value=f'Contingência ({budget.contingency_pct}%):').font = total_font
        ws.cell(row=totals_row + 1, column=6, value=float(budget.total_contingency)).font = total_font
        ws.cell(row=totals_row + 1, column=6).number_format = '#,##0.00 CVE'
        ws.cell(row=totals_row + 1, column=6).fill = total_fill
        
        ws.cell(row=totals_row + 2, column=1, value='TOTAL:').font = Font(bold=True, size=12)
        ws.cell(row=totals_row + 2, column=6, value=float(budget.grand_total)).font = Font(bold=True, size=12)
        ws.cell(row=totals_row + 2, column=6).number_format = '#,##0.00 CVE'
        ws.cell(row=totals_row + 2, column=6).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        # Larguras das colunas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        if include_notes:
            ws.column_dimensions['G'].width = 30
        
        # Salvar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="orcamento_{budget.name.replace(" ", "_")}.xlsx"'
        return response


def generate_budget_pdf_response(budget, include_notes=True):
    """Função auxiliar para gerar resposta PDF."""
    renderer = BudgetPDFRenderer()
    return renderer.render(budget, include_notes)


def generate_budget_excel_response(budget, include_notes=True):
    """Função auxiliar para gerar resposta Excel."""
    renderer = BudgetExcelRenderer()
    return renderer.render(budget, include_notes)

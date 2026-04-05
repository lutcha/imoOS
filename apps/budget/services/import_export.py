"""
Import/Export — Excel import/export para orçamentos.

Suporta:
- Importação de BOQ (Bill of Quantities) em Excel
- Exportação de orçamentos para Excel
- Validação de dados durante importação
"""
import io
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Tuple, Optional
from datetime import datetime

from django.core.files.uploadedfile import UploadedFile

from apps.budget.models import SimpleBudget, BudgetItem, LocalPriceItem


class ExcelImporter:
    """
    Importador de orçamentos a partir de ficheiros Excel.
    
    Formato esperado:
    - Coluna A: Nº Linha
    - Coluna B: Categoria (MATERIALS/LABOR/EQUIPMENT/SERVICES)
    - Coluna C: Descrição
    - Coluna D: Quantidade
    - Coluna E: Unidade
    - Coluna F: Preço Unitário
    - Coluna G: Observações (opcional)
    """
    
    REQUIRED_COLUMNS = ['line_number', 'category', 'description', 'quantity', 'unit', 'unit_price']
    
    CATEGORY_MAPPING = {
        'material': 'MATERIALS',
        'materiais': 'MATERIALS',
        'materiais de construção': 'MATERIALS',
        'mat': 'MATERIALS',
        'labor': 'LABOR',
        'mão-de-obra': 'LABOR',
        'mao-de-obra': 'LABOR',
        'mão de obra': 'LABOR',
        'obra': 'LABOR',
        'equipment': 'EQUIPMENT',
        'equipamentos': 'EQUIPMENT',
        'equip': 'EQUIPMENT',
        'services': 'SERVICES',
        'serviços': 'SERVICES',
        'servicos': 'SERVICES',
    }
    
    UNIT_MAPPING = {
        'un': 'UN',
        'unidade': 'UN',
        'm2': 'M2',
        'm²': 'M2',
        'metro quadrado': 'M2',
        'm3': 'M3',
        'm³': 'M3',
        'metro cúbico': 'M3',
        'metro cubico': 'M3',
        'kg': 'KG',
        'quilograma': 'KG',
        'hr': 'HR',
        'hora': 'HR',
        'day': 'DAY',
        'dia': 'DAY',
        'saco': 'SACO',
        'l': 'L',
        'litro': 'L',
        'ml': 'ML',
        'metro linear': 'ML',
        'kit': 'KIT',
    }
    
    def __init__(self, budget: SimpleBudget):
        self.budget = budget
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
        self.imported_count = 0
    
    def import_from_file(self, file: UploadedFile) -> Dict:
        """
        Importar itens de um ficheiro Excel.
        
        Args:
            file: Ficheiro Excel (.xlsx ou .xls)
            
        Returns:
            Dicionário com resultado da importação
        """
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        
        try:
            import openpyxl
        except ImportError:
            return {
                'success': False,
                'errors': [{'message': 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl'}],
                'imported_count': 0
            }
        
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
        except Exception as e:
            return {
                'success': False,
                'errors': [{'message': f'Erro ao ler ficheiro Excel: {str(e)}'}],
                'imported_count': 0
            }
        
        # Detectar cabeçalhos na primeira linha
        headers = self._detect_headers(sheet)
        if not headers:
            return {
                'success': False,
                'errors': [{'message': 'Não foi possível detectar cabeçalhos. Formato esperado: Nº, Categoria, Descrição, Qtd, Unidade, Preço'}],
                'imported_count': 0
            }
        
        # Processar linhas
        row_number = 2  # Começar após cabeçalho
        for row in sheet.iter_rows(min_row=2, values_only=True):
            result = self._process_row(row, headers, row_number)
            if result['success']:
                self.imported_count += 1
            row_number += 1
        
        # Recalcular totais do orçamento
        if self.imported_count > 0:
            self.budget.recalculate_totals()
        
        return {
            'success': len(self.errors) == 0,
            'imported_count': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def _detect_headers(self, sheet) -> Optional[Dict]:
        """Detectar e mapear cabeçalhos da planilha."""
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        if not first_row:
            return None
        
        headers = {}
        header_keywords = {
            'line_number': ['nº', 'num', 'linha', 'line', 'item', '#'],
            'category': ['categoria', 'category', 'tipo', 'type'],
            'description': ['descrição', 'descricao', 'description', 'desc', 'item'],
            'quantity': ['quantidade', 'qtd', 'quantity', 'qty'],
            'unit': ['unidade', 'unit', 'und', 'uni'],
            'unit_price': ['preço unit', 'preco unit', 'unit price', 'p.unit', 'unitário'],
            'notes': ['obs', 'observações', 'observacoes', 'notes', 'notas'],
        }
        
        for idx, cell_value in enumerate(first_row):
            if not cell_value:
                continue
            
            cell_str = str(cell_value).lower().strip()
            
            for field, keywords in header_keywords.items():
                if any(kw in cell_str for kw in keywords):
                    headers[field] = idx
                    break
        
        # Verificar se temos campos obrigatórios mínimos
        if 'description' in headers and 'quantity' in headers:
            # Definir defaults para campos faltantes
            if 'line_number' not in headers:
                headers['line_number'] = None  # Auto-increment
            if 'category' not in headers:
                headers['category'] = None  # Default MATERIALS
            if 'unit' not in headers:
                headers['unit'] = None  # Default UN
            if 'unit_price' not in headers:
                headers['unit_price'] = None  # Default 0
            
            return headers
        
        return None
    
    def _process_row(self, row: Tuple, headers: Dict, row_number: int) -> Dict:
        """Processar uma linha da planilha."""
        result = {'success': False, 'row': row_number}
        
        # Extrair valores
        line_number = self._get_cell_value(row, headers.get('line_number'), row_number - 1)
        category = self._normalize_category(self._get_cell_value(row, headers.get('category'), 'MATERIALS'))
        description = self._get_cell_value(row, headers.get('description'), '').strip()
        quantity = self._parse_decimal(self._get_cell_value(row, headers.get('quantity'), 0))
        unit = self._normalize_unit(self._get_cell_value(row, headers.get('unit'), 'UN'))
        unit_price = self._parse_decimal(self._get_cell_value(row, headers.get('unit_price'), 0))
        notes = self._get_cell_value(row, headers.get('notes'), '')
        
        # Validações
        if not description:
            self.errors.append({
                'row': row_number,
                'message': 'Descrição é obrigatória'
            })
            return result
        
        if quantity is None or quantity < 0:
            self.errors.append({
                'row': row_number,
                'message': f'Quantidade inválida: {row[headers.get("quantity")] if headers.get("quantity") is not None else ""}'
            })
            return result
        
        if unit_price is None or unit_price < 0:
            self.errors.append({
                'row': row_number,
                'message': f'Preço unitário inválido: {row[headers.get("unit_price")] if headers.get("unit_price") is not None else ""}'
            })
            return result
        
        # Tentar encontrar preço na base de dados se não especificado
        if unit_price == 0:
            db_price = self._lookup_price(description, category, unit)
            if db_price:
                unit_price = db_price
                self.warnings.append({
                    'row': row_number,
                    'message': f'Preço unitário obtido da base de dados: {db_price}'
                })
        
        # Criar item
        try:
            BudgetItem.objects.create(
                budget=self.budget,
                line_number=line_number if isinstance(line_number, int) else row_number - 1,
                category=category,
                description=description,
                quantity=quantity,
                unit=unit,
                unit_price=unit_price,
                price_source=BudgetItem.PRICE_SOURCE_DATABASE if unit_price > 0 else BudgetItem.PRICE_SOURCE_MANUAL,
                notes=notes
            )
            result['success'] = True
        except Exception as e:
            self.errors.append({
                'row': row_number,
                'message': f'Erro ao criar item: {str(e)}'
            })
        
        return result
    
    def _get_cell_value(self, row: Tuple, index: Optional[int], default=None):
        """Obter valor de uma célula com segurança."""
        if index is None or index >= len(row):
            return default
        value = row[index]
        if value is None:
            return default
        return value
    
    def _normalize_category(self, value: str) -> str:
        """Normalizar categoria para valor válido."""
        if not value:
            return 'MATERIALS'
        
        normalized = str(value).lower().strip()
        return self.CATEGORY_MAPPING.get(normalized, 'MATERIALS')
    
    def _normalize_unit(self, value: str) -> str:
        """Normalizar unidade para valor válido."""
        if not value:
            return 'UN'
        
        normalized = str(value).lower().strip()
        return self.UNIT_MAPPING.get(normalized, 'UN')
    
    def _parse_decimal(self, value) -> Optional[Decimal]:
        """Converter valor para Decimal."""
        if value is None:
            return Decimal('0')
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        try:
            # Remover separadores de milhar e converter decimal
            cleaned = str(value).replace(',', '.').replace(' ', '')
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return None
    
    def _lookup_price(self, description: str, category: str, unit: str) -> Optional[Decimal]:
        """Buscar preço na base de dados local."""
        from apps.budget.services import PriceEngine
        engine = PriceEngine()
        
        suggestion = engine.suggest_price(
            item_name=description,
            island=self.budget.island,
            category=category,
            unit=unit
        )
        
        if suggestion['suggested_price']:
            return Decimal(str(suggestion['suggested_price']))
        
        return None


class ExcelExporter:
    """
    Exportador de orçamentos para Excel.
    
    Gera planilhas formatadas com:
    - Cabeçalhos estilizados
    - Fórmulas para totais
    - Resumo por categoria
    """
    
    def __init__(self, budget: SimpleBudget):
        self.budget = budget
    
    def export_to_bytes(self) -> bytes:
        """
        Exportar orçamento para bytes Excel.
        
        Returns:
            Bytes do ficheiro Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError('Biblioteca openpyxl não instalada. Execute: pip install openpyxl')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Orçamento'
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(size=11)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        sheet['A1'] = f'Orçamento: {self.budget.name}'
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:G1')
        
        sheet['A2'] = f'Projecto: {self.budget.project.name}'
        sheet['A2'].font = subtitle_font
        sheet.merge_cells('A2:G2')
        
        sheet['A3'] = f'Ilha: {self.budget.get_island_display()} | Versão: {self.budget.version} | Data: {datetime.now().strftime("%d/%m/%Y")}'
        sheet['A3'].font = subtitle_font
        sheet.merge_cells('A3:G3')
        
        # Cabeçalhos (linha 5)
        headers = ['Nº', 'Categoria', 'Descrição', 'Quantidade', 'Unidade', 'Preço Unit.', 'Total', 'Obs.']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Items
        row = 6
        for item in self.budget.items.all():
            sheet.cell(row=row, column=1, value=item.line_number).border = border
            sheet.cell(row=row, column=2, value=item.get_category_display()).border = border
            sheet.cell(row=row, column=3, value=item.description).border = border
            sheet.cell(row=row, column=4, value=float(item.quantity)).border = border
            sheet.cell(row=row, column=5, value=item.get_unit_display()).border = border
            sheet.cell(row=row, column=6, value=float(item.unit_price)).number_format = '#,##0.00'
            sheet.cell(row=row, column=6).border = border
            sheet.cell(row=row, column=7, value=float(item.total)).number_format = '#,##0.00'
            sheet.cell(row=row, column=7).border = border
            sheet.cell(row=row, column=8, value=item.notes).border = border
            row += 1
        
        # Linha em branco
        row += 1
        
        # Totais
        totals_start = row
        self._add_total_row(sheet, row, 'Total Materiais:', float(self.budget.total_materials), border)
        row += 1
        self._add_total_row(sheet, row, 'Total Mão-de-Obra:', float(self.budget.total_labor), border)
        row += 1
        self._add_total_row(sheet, row, 'Total Equipamentos:', float(self.budget.total_equipment), border)
        row += 1
        self._add_total_row(sheet, row, 'Total Serviços:', float(self.budget.total_services), border)
        row += 1
        self._add_total_row(sheet, row, 'Subtotal:', float(self.budget.subtotal), border, bold=True)
        row += 1
        self._add_total_row(sheet, row, f'Contingência ({self.budget.contingency_pct}%):', float(self.budget.total_contingency), border)
        row += 1
        
        # Total geral destacado
        total_fill = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
        sheet.cell(row=row, column=6, value='TOTAL GERAL:').font = Font(bold=True, size=12)
        sheet.cell(row=row, column=6).fill = total_fill
        sheet.cell(row=row, column=6).border = border
        sheet.cell(row=row, column=7, value=float(self.budget.grand_total)).number_format = '#,##0.00'
        sheet.cell(row=row, column=7).font = Font(bold=True, size=12)
        sheet.cell(row=row, column=7).fill = total_fill
        sheet.cell(row=row, column=7).border = border
        
        # Ajustar larguras das colunas
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 45
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 14
        sheet.column_dimensions['H'].width = 25
        
        # Congelar painel no cabeçalho
        sheet.freeze_panes = 'A6'
        
        # Guardar em bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _add_total_row(self, sheet, row: int, label: str, value: float, border, bold: bool = False):
        """Adicionar linha de total."""
        sheet.cell(row=row, column=6, value=label)
        sheet.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        if bold:
            sheet.cell(row=row, column=6).font = Font(bold=True)
        sheet.cell(row=row, column=6).border = border
        
        sheet.cell(row=row, column=7, value=value)
        sheet.cell(row=row, column=7).number_format = '#,##0.00'
        if bold:
            sheet.cell(row=row, column=7).font = Font(bold=True)
        sheet.cell(row=row, column=7).border = border
